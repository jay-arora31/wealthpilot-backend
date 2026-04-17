import io
from decimal import Decimal, InvalidOperation

import logfire
from fastapi import UploadFile

from app.agents.column_mapping import (
    ColumnMapping,
    column_mapping_agent,
    mapping_to_display,
    normalize_mapping,
    review_mapping,
)
from app.core import jobs as job_store
from app.repositories.account_repo import AccountRepository
from app.repositories.bank_detail_repo import BankDetailRepository
from app.repositories.household_repo import HouseholdRepository
from app.repositories.member_repo import MemberRepository
from app.schemas.account import AccountCreate, OwnershipCreate
from app.schemas.bank_detail import BankDetailCreate
from app.schemas.household import HouseholdCreate
from app.schemas.member import MemberCreate
from app.services.conflict_service import ConflictService


def _safe_decimal(value) -> Decimal | None:
    if value is None:
        return None
    try:
        d = Decimal(str(value).replace(",", "").replace("$", "").strip())
        # Treat 0 as missing data — Excel often has 0.0 as a placeholder
        return d if d != 0 else None
    except InvalidOperation:
        return None


def _safe_percentage(value) -> Decimal | None:
    """Parse ownership percentage — must be between 0 and 100. Returns None if invalid."""
    d = _safe_decimal(value)
    if d is None:
        return None
    if d < 0 or d > 100:
        return None
    return d


def _safe_str(value) -> str | None:
    if value is None:
        return None
    # Excel returns numeric cells as floats (e.g. 9977221122.0).
    # Convert whole-number floats to int first so the string has no trailing ".0".
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    # Strip time portion from datetime objects (e.g. 1968-11-01 00:00:00 → 1968-11-01)
    if hasattr(value, "date"):
        value = value.date()
    s = str(value).strip()
    return s if s else None


def _safe_dob(value) -> str | None:
    """Parse date of birth, handling datetime objects AND integer dates like 12251969 (MMDDYYYY)."""
    if value is None:
        return None
    # datetime object from openpyxl
    if hasattr(value, "date"):
        return str(value.date())
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, int):
        s = str(value)
        # Try MMDDYYYY (8 digits)
        if len(s) == 8:
            try:
                from datetime import date
                return str(date(int(s[4:]), int(s[:2]), int(s[2:4])))
            except Exception:
                pass
        # Try YYYYMMDD
        if len(s) == 8:
            try:
                from datetime import date
                return str(date(int(s[:4]), int(s[4:6]), int(s[6:])))
            except Exception:
                pass
        return s
    s = str(value).strip()
    return s if s else None


def _safe_tax_bracket(value) -> str | None:
    """Normalise tax bracket: 0.25 → '25%', 'Highest' → 'Highest', '25%' → '25%'."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, (int, float, Decimal)):
        try:
            d = Decimal(str(value))
            # Values like 0.25 represent 25%; values like 25 represent 25%
            pct = d * 100 if d <= 1 else d
            # Use float formatting so trailing zeros are dropped (e.g. 25.00 → "25")
            return f"{float(pct):g}%"
        except Exception:
            pass
    s = str(value).strip()
    return s if s else None


def _get_col(row: list, idx: int | None):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _resolve_member_name(row: list, mapping) -> str | None:
    """Return member full name from a dedicated column or by combining first + last."""
    if mapping.member_name is not None:
        return _safe_str(_get_col(row, mapping.member_name))
    first = _safe_str(_get_col(row, mapping.member_first_name))
    last = _safe_str(_get_col(row, mapping.member_last_name))
    parts = [p for p in [first, last] if p]
    return " ".join(parts) if parts else None


class ExcelService:
    def __init__(
        self,
        household_repo: HouseholdRepository,
        member_repo: MemberRepository,
        account_repo: AccountRepository,
        bank_detail_repo: BankDetailRepository,
        conflict_service: ConflictService,
    ) -> None:
        self.household_repo = household_repo
        self.member_repo = member_repo
        self.account_repo = account_repo
        self.bank_detail_repo = bank_detail_repo
        self.conflict_service = conflict_service

    @logfire.instrument("excel.process_excel", extract_args=False)
    async def process_excel(self, file: UploadFile, job_id: str | None = None) -> dict:
        import openpyxl

        def _step(msg: str) -> None:
            if job_id:
                job_store.append_step(job_id, msg)

        logfire.info("excel.upload_received", filename=file.filename, job_id=job_id)
        _step("Reading file…")
        content = await file.read()

        _step("Opening workbook…")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        created = 0
        enriched = 0
        conflicts_count = 0
        enriched_names: set[str] = set()
        created_names: set[str] = set()
        all_mappings: list[dict] = []

        sheets = wb.sheetnames
        _step(f"Found {len(sheets)} sheet{'s' if len(sheets) != 1 else ''}: {', '.join(sheets)}")

        for sheet_name in sheets:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                _step(f"Sheet '{sheet_name}' is empty — skipping")
                continue

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            if not any(headers):
                _step(f"Sheet '{sheet_name}' has no headers — skipping")
                continue

            _step(f"Sheet '{sheet_name}': {len(rows) - 1} data rows detected")
            _step(f"Running AI column mapping on '{sheet_name}'…")

            sample_rows = [list(r) for r in rows[1:6]]
            with logfire.span("excel.column_mapping_ai", sheet=sheet_name):
                result = await column_mapping_agent.run(
                    f"Headers: {headers}\nSample rows: {sample_rows}"
                )
            proposed: ColumnMapping = result.output

            _step(f"Reviewing AI mapping on '{sheet_name}'…")
            review_rows = [list(r) for r in rows[1:11]]
            with logfire.span("excel.column_mapping_review", sheet=sheet_name):
                reviewed: ColumnMapping = await review_mapping(proposed, headers, review_rows)
            mapping: ColumnMapping = normalize_mapping(reviewed)

            display = mapping_to_display(mapping, headers)
            all_mappings.append({"sheet": sheet_name, "mappings": display})
            mapped_pairs = [f"{m['field']}→\"{m['header']}\"" for m in display]
            _step(f"Mapped {len(mapped_pairs)} columns: {', '.join(mapped_pairs[:8])}{'…' if len(mapped_pairs) > 8 else ''}")

            # Pre-scan all rows to aggregate the best non-null household-level
            # fields across every row that belongs to the same household.
            # This handles spreadsheets where risk/horizon etc. appear on a
            # sub-row rather than the first row for a household.
            hh_aggregated: dict[str, dict] = {}
            agg_fields = [
                "income", "net_worth", "liquid_net_worth", "expense_range",
                "tax_bracket", "risk_tolerance", "time_horizon", "goals", "preferences",
            ]
            for raw_row in rows[1:]:
                row = list(raw_row)
                if not any(v is not None for v in row):
                    continue
                hh_name = _safe_str(_get_col(row, mapping.household_name))
                if not hh_name:
                    continue
                hh_name = hh_name.strip()
                if hh_name not in hh_aggregated:
                    hh_aggregated[hh_name] = {f: None for f in agg_fields}
                agg = hh_aggregated[hh_name]
                if agg["income"] is None:
                    agg["income"] = _safe_decimal(_get_col(row, mapping.income))
                if agg["net_worth"] is None:
                    agg["net_worth"] = _safe_decimal(_get_col(row, mapping.net_worth))
                if agg["liquid_net_worth"] is None:
                    agg["liquid_net_worth"] = _safe_decimal(_get_col(row, mapping.liquid_net_worth))
                if agg["expense_range"] is None:
                    agg["expense_range"] = _safe_str(_get_col(row, mapping.expense_range))
                if agg["tax_bracket"] is None:
                    agg["tax_bracket"] = _safe_tax_bracket(_get_col(row, mapping.tax_bracket))
                if agg["risk_tolerance"] is None:
                    agg["risk_tolerance"] = _safe_str(_get_col(row, mapping.risk_tolerance))
                if agg["time_horizon"] is None:
                    agg["time_horizon"] = _safe_str(_get_col(row, mapping.time_horizon))
                if agg["goals"] is None:
                    agg["goals"] = _safe_str(_get_col(row, mapping.goals))
                if agg["preferences"] is None:
                    agg["preferences"] = _safe_str(_get_col(row, mapping.preferences))

            for raw_row in rows[1:]:
                row = list(raw_row)
                if not any(v is not None for v in row):
                    continue

                household_name = _safe_str(_get_col(row, mapping.household_name))
                if not household_name:
                    continue
                household_name = household_name.strip()

                # Everything written for this row goes into a single transaction.
                # Repos flush (to get back generated IDs) but don't commit; we
                # commit once at the bottom of the row. This collapses ~14
                # per-row transactions into 1, removing a matching number of
                # pgbouncer reconnects per Excel row.
                db_session = self.household_repo.db

                agg = hh_aggregated.get(household_name, {})
                household_data = HouseholdCreate(
                    name=household_name,
                    income=agg.get("income"),
                    net_worth=agg.get("net_worth"),
                    liquid_net_worth=agg.get("liquid_net_worth"),
                    expense_range=agg.get("expense_range"),
                    tax_bracket=agg.get("tax_bracket"),
                    risk_tolerance=agg.get("risk_tolerance"),
                    time_horizon=agg.get("time_horizon"),
                    goals=agg.get("goals"),
                    preferences=agg.get("preferences"),
                )

                existing = await self.household_repo.find_by_name(household_name)

                if existing:
                    numeric_fields = {"income", "net_worth", "liquid_net_worth"}
                    financial_fields = [
                        "income", "net_worth", "liquid_net_worth", "expense_range",
                        "tax_bracket", "risk_tolerance", "time_horizon", "goals", "preferences",
                    ]
                    row_conflicts = 0
                    for field in financial_fields:
                        incoming_val = getattr(household_data, field)
                        existing_val = getattr(existing, field)
                        if incoming_val is None:
                            continue
                        # Compare numeric fields as Decimal to avoid false conflicts
                        # from formatting differences (e.g. 150000.00 vs 150000.0)
                        if field in numeric_fields:
                            try:
                                if Decimal(str(incoming_val)) == Decimal(str(existing_val or 0)):
                                    continue
                            except Exception:
                                pass
                        else:
                            if str(incoming_val) == str(existing_val or ""):
                                continue
                        await self.conflict_service.create_conflict(
                            household_id=existing.id,
                            field=field,
                            existing=str(existing_val) if existing_val is not None else None,
                            incoming=str(incoming_val),
                            source="excel",
                            commit=False,
                        )
                        conflicts_count += 1
                        row_conflicts += 1

                    member_name = _resolve_member_name(row, mapping)
                    enrich_member = None
                    if member_name:
                        member_dob = _safe_dob(_get_col(row, mapping.member_dob))
                        existing_member = await self.member_repo.find_by_name_in_household(
                            existing.id, member_name, date_of_birth=member_dob
                        )
                        if not existing_member:
                            enrich_member = await self.member_repo.create(
                                existing.id,
                                MemberCreate(
                                    name=member_name,
                                    date_of_birth=member_dob,
                                    email=_safe_str(_get_col(row, mapping.member_email)),
                                    phone=_safe_str(_get_col(row, mapping.member_phone)),
                                    member_relationship=_safe_str(_get_col(row, mapping.member_relationship)),
                                    address=_safe_str(_get_col(row, mapping.member_address)),
                                ),
                                commit=False,
                            )
                        else:
                            # Backfill missing DOB if this row provides one
                            if member_dob and existing_member.date_of_birth is None:
                                await self.member_repo.update_dob(
                                    existing_member.id, member_dob, commit=False
                                )
                            enrich_member = existing_member

                    # Add accounts that don't already exist on the household
                    enrich_acct_number = _safe_str(_get_col(row, mapping.account_number))
                    enrich_acct_type = _safe_str(_get_col(row, mapping.account_type))
                    if enrich_acct_number or enrich_acct_type:
                        existing_acct = None
                        if enrich_acct_number:
                            existing_acct = await self.account_repo.find_by_account_number(
                                existing.id, enrich_acct_number
                            )
                        elif enrich_acct_type:
                            existing_acct = await self.account_repo.find_by_type_in_household(
                                existing.id, enrich_acct_type
                            )
                        if not existing_acct:
                            ownership_pct = _safe_percentage(_get_col(row, mapping.ownership_percentage))
                            ownerships = (
                                [OwnershipCreate(member_id=enrich_member.id, ownership_percentage=ownership_pct)]
                                if ownership_pct and enrich_member
                                else []
                            )
                            await self.account_repo.create(
                                existing.id,
                                AccountCreate(
                                    account_number=enrich_acct_number,
                                    custodian=_safe_str(_get_col(row, mapping.custodian)),
                                    account_type=enrich_acct_type,
                                    account_value=_safe_decimal(_get_col(row, mapping.account_value)),
                                    ownerships=ownerships,
                                ),
                                commit=False,
                            )

                    if household_name not in enriched_names:
                        enriched_names.add(household_name)
                        enriched += 1
                        conflict_note = f" ({row_conflicts} conflicts)" if row_conflicts else ""
                        _step(f"Enriched: '{household_name}'{conflict_note}")
                else:
                    new_household = await self.household_repo.create(
                        household_data, commit=False
                    )

                    member_name = _resolve_member_name(row, mapping)
                    member = None
                    if member_name:
                        member_dob = _safe_dob(_get_col(row, mapping.member_dob))
                        existing_new_member = await self.member_repo.find_by_name_in_household(
                            new_household.id, member_name, date_of_birth=member_dob
                        )
                        if not existing_new_member:
                            member = await self.member_repo.create(
                                new_household.id,
                                MemberCreate(
                                    name=member_name,
                                    date_of_birth=member_dob,
                                    email=_safe_str(_get_col(row, mapping.member_email)),
                                    phone=_safe_str(_get_col(row, mapping.member_phone)),
                                    member_relationship=_safe_str(_get_col(row, mapping.member_relationship)),
                                    address=_safe_str(_get_col(row, mapping.member_address)),
                                ),
                                commit=False,
                            )
                        else:
                            if member_dob and existing_new_member.date_of_birth is None:
                                await self.member_repo.update_dob(
                                    existing_new_member.id, member_dob, commit=False
                                )
                            member = existing_new_member
                        account_number = _safe_str(_get_col(row, mapping.account_number))
                        account_type = _safe_str(_get_col(row, mapping.account_type))
                        # Create account if we have either a number or a type
                        if account_number or account_type:
                            # Deduplicate: skip if same account_number already exists,
                            # or (when no number) same account_type already exists
                            existing_acct = None
                            if account_number:
                                existing_acct = await self.account_repo.find_by_account_number(
                                    new_household.id, account_number
                                )
                            elif account_type:
                                existing_acct = await self.account_repo.find_by_type_in_household(
                                    new_household.id, account_type
                                )
                            if not existing_acct:
                                ownership_pct = _safe_percentage(_get_col(row, mapping.ownership_percentage))
                                ownerships = (
                                    [OwnershipCreate(member_id=member.id, ownership_percentage=ownership_pct)]
                                    if ownership_pct
                                    else []
                                )
                                await self.account_repo.create(
                                    new_household.id,
                                    AccountCreate(
                                        account_number=account_number,
                                        custodian=_safe_str(_get_col(row, mapping.custodian)),
                                        account_type=account_type,
                                        account_value=_safe_decimal(_get_col(row, mapping.account_value)),
                                        ownerships=ownerships,
                                    ),
                                    commit=False,
                                )

                    bank_name = _safe_str(_get_col(row, mapping.bank_name))
                    if bank_name:
                        await self.bank_detail_repo.create(
                            new_household.id,
                            BankDetailCreate(
                                bank_name=bank_name,
                                account_number=_safe_str(_get_col(row, mapping.bank_account_number)),
                                routing_number=_safe_str(_get_col(row, mapping.routing_number)),
                            ),
                            commit=False,
                        )

                    if household_name not in created_names:
                        created_names.add(household_name)
                        created += 1
                        _step(f"Created: '{household_name}'")

                # Commit everything accumulated for this row in a single
                # transaction. On failure, roll back the row and keep going so
                # one bad row doesn't abort the whole upload.
                try:
                    await db_session.commit()
                except Exception as exc:
                    await db_session.rollback()
                    logfire.warning(
                        "excel.row_commit_failed",
                        sheet=sheet_name,
                        household=household_name,
                        error=str(exc),
                    )
                    _step(f"Row failed for '{household_name}' — rolled back ({exc})")

        summary = {
            "created": created,
            "enriched": enriched,
            "conflicts": conflicts_count,
            "column_mappings": all_mappings,
        }
        logfire.info(
            "excel.processing_complete",
            created=created,
            enriched=enriched,
            conflicts=conflicts_count,
        )
        _step(
            f"Done — {created} created, {enriched} enriched, {conflicts_count} conflict{'s' if conflicts_count != 1 else ''}"
        )
        return summary
